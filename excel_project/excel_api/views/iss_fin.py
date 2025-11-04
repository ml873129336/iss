from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser,FormParser
import pandas as pd
import io,os
from openpyxl import load_workbook
import tempfile
from django.http import FileResponse,Http404
from rest_framework.views import APIView
from rest_framework import status
from django.conf import settings
from utils import mail_utils
import platform


class Iss_Fin_solve_excel(APIView):
    parser_classes = (MultiPartParser, FormParser)



    def post(self, request):

        self.df1 = None
        self.df2 = None
        self.df3 = None
        self.df4 = None
        files = request.FILES.getlist("files")

        try:
            self.read_files(files)

            # non_empty_count = sum(df is not None and not df.empty for df in success_df)
            res = self.solve_excel_data()

            return res
            # return Response({
            #     'message': '上传成功',
            #     'count': non_empty_count,
            #     'files': non_empty_count
            # }, status=status.HTTP_201_CREATED)

        except Exception as e:
            print(e)
            return Response({
                'message': f'上传失败 {e}',
                'count': 0,
                'files': 0
            }, status=status.HTTP_201_CREATED)


    def get(self, request):
        self.df1 = None
        self.df2 = None
        self.df3 = None
        self.df4 = None
        try:
            files = mail_utils.check_email("INV")
            print(files)
            # self.read_files(files)
            # # non_empty_count = sum(df is not None and not df.empty for df in success_df)
            # res = self.solve_excel_data()
            res = Response({
                'message': f'1',
                'count': 0,
                'files': 0
            }, status=status.HTTP_201_CREATED)
            return res
            # return Response({
            #     'message': '上传成功',
            #     'count': non_empty_count,
            #     'files': non_empty_count
            # }, status=status.HTTP_201_CREATED)

        except Exception as e:
            print(e)
            return Response({
                'message': f'上传失败 {e}',
                'count': 0,
                'files': 0
            }, status=status.HTTP_201_CREATED)

    def read_files(self,files):

        for file in files:

            if isinstance(file, str):
                filename = file
                content = None
            else:
                filename = file.name
                content = file.read()
                file.seek(0)

            if "eInvoice" in file.name:
                # self.df1 = pd.read_excel(io.BytesIO(file.read()))
                if content:
                    self.df1 = pd.read_excel(io.BytesIO(content))
                else:
                    self.df1 = pd.read_excel(file)
            elif "Invoices" in file.name:
                # self.df2 = pd.read_csv(io.BytesIO(file.read()), usecols=["Invoice No.", "Created By"],
                #                        encoding="utf-8-sig")
                if content:
                    self.df2 = pd.read_csv(io.BytesIO(content), usecols=["Invoice No.", "Created By"],
                                           encoding="utf-8-sig")
                else:
                    self.df2 = pd.read_csv(file, usecols=["Invoice No.", "Created By"], encoding="utf-8-sig")

            elif file.name.startswith("Contacts"):
                self.df4 = pd.read_excel(io.BytesIO(content),usecols=["Foreign Name", "Tax ID"])
                print(self.df4)
            else:
                sheets_to_read = ['海运部', '深圳办', '宁波办', '空运部', '项目部','税率6%']
                # sheets_to_read = ['空运部']
                # file.bytes = file.read()
                # dfs = [pd.read_excel(io.BytesIO(file.bytes), sheet_name=s, skiprows=1,
                #                      usecols=["INVPVGNo.\n不要有空格", "开票商品名称", "备注"]) for s in sheets_to_read]
                if content:
                    dfs = [
                        pd.read_excel(io.BytesIO(content), sheet_name=s, skiprows=1,
                                      usecols=["INVPVGNo.\n不要有空格", "开票商品名称", "备注"])
                        for s in sheets_to_read
                    ]
                else:
                    dfs = [
                        pd.read_excel(file, sheet_name=s, skiprows=1,
                                      usecols=["INVPVGNo.\n不要有空格", "开票商品名称", "备注"])
                        for s in sheets_to_read
                    ]
                self.df3 = pd.concat(dfs, ignore_index=True)
                self.df3 = self.df3.replace('_x000D_', '', regex=True)
                self.df3['INVPVGNo.\n不要有空格'] = self.df3['INVPVGNo.\n不要有空格'].str.replace(r'\s+', '', regex=True)
                # self.check_both_loaded()

    def update_row(self,row):
        if row['Invoice type'] == "02" :
            return "国际货物运输" + str(row['开票商品名称'])
        return str(row['开票商品名称'])

    def solve_excel_data(self):
        if self.df1 is None or self.df2 is None:
            return

        df1 = self.df1.drop(self.df1.index[0]).reset_index(drop=True)

        df = df1.fillna('')
        cols = ['Amount', 'Exchange rate']
        for col in cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        df['sum1'] = df['Amount'] * df['Exchange rate']
        df['sum'] = df.groupby(['Invoice number'])['sum1'].transform('sum')
        df['sum_yuan']=df.groupby(['Invoice number'])['Amount'].transform('sum')

        last_col_data = df.iloc[:, -2]
        df.loc[:, 'Unit price'] = last_col_data
        # df.loc[:, 'Amount'] = last_col_data
        df = df.rename(columns={"Invoice number": "Invoice No."})
        df.drop(columns=['sum1'], inplace=True)

        self.df2["email"] = self.df2["Created By"].apply(self.name_to_email)

        inv_to_email = pd.Series(self.df2["email"].values, index=self.df2["Invoice No."]).to_dict()
        df["Email address"] = df["Invoice No."].map(inv_to_email)
        df = df.drop_duplicates(subset=['Invoice No.'])


        if self.df3 is not None and not self.df3.empty:

            df = df[df['Invoice No.'].isin(self.df3['INVPVGNo.\n不要有空格'])]
            df_merged = df.merge(
                self.df3[['INVPVGNo.\n不要有空格', '开票商品名称', '备注']],
                left_on='Invoice No.',
                right_on='INVPVGNo.\n不要有空格',
                how='left',
            )

            print(df_merged.columns)
            # print(df_merged['Tax rate'])



            df_merged['Commodity name'] = df_merged.apply(self.update_row, axis=1)


            df_merged['备注'] = df_merged['备注'].fillna('').astype(str)
            df_merged['Issuing note'] = df_merged['Issuing note'] + "\n" + df_merged['备注']

            df = df_merged.drop(columns=['开票商品名称', '备注'])
            print(df)
            if self.df4 is not None and not self.df4.empty:
                self.df4 = self.df4.drop_duplicates(subset=['Foreign Name'], keep='first')
                df = df.merge(self.df4, left_on='Buyer name',right_on="Foreign Name", how='left')
                print(df)
                df['Buyer tax number'] = df['Buyer tax number'].mask(df['Buyer tax number'] == '', df['Tax ID'])
                df = df.drop(columns=['Tax ID'])

            self.filter_df = df

            self.complete_template_data(self.filter_df)
        return self.download_file("output.xlsx")


    def complete_template_data(self, df):

        if platform.system() == 'Windows':
            # Windows 系统
            template_path = r'D:\project\excelapp\excel_project\excel_api\templates\excel_api\template.xlsx'
        else:
            # Linux / macOS
            template_path = os.path.normpath(
                os.path.join(
                    settings.BASE_DIR,
                    'excel_api',
                    'templates',
                    'excel_api',
                    'template.xlsx'  # 注意 Linux 文件名大小写敏感
                )
            )




        # template_path = os.path.normpath(os.path.join(settings.BASE_DIR, 'excel_api', 'templates', 'excel_api',  '模版.xlsx'))
        # template_path = r'D:\project\excelapp\excel_project\excel_api\templates\excel_api\template.xlsx'
        df["zhence"] = ''
        df["biaoshi"] = ''
        df.loc[df['Invoice type'] == '02', 'zhence'] = '免税'
        df.loc[df['Invoice type'] == '02', 'biaoshi'] = 1
        wb = load_workbook(template_path)

        ws = wb.active  # 如果有多个 sheet，可以 ws = wb["Sheet1"]

        template_columns = [ws.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)]


        start_row = 4
        col_map = {
            "开票单号*": "Invoice No.",
            "购方名称*": "Buyer name",
            "购方税号": "Buyer tax number",
            "商品名称*": "Commodity name",
            "单价": "Unit price",
            "金额": "sum",
            "发票备注": "Issuing note",
            "邮箱地址": "Email address",
            "税率": "Tax rate",
            "发票种类*": "Invoice type",
            "零税率标识": "biaoshi",
            "优惠政策名称": "zhence",
            "原金额": "Amount"

        }


        for row_idx, row_data in enumerate(df.to_dict(orient="records"), start=start_row):
            for col_idx, template_col in enumerate(template_columns, start=1):
                if template_col == "税收分类编码":
                    ws.cell(row=row_idx, column=col_idx, value="3040802010200000000")
                if template_col == "数量":
                    ws.cell(row=row_idx, column=col_idx, value="1.00")
                if template_col == "清单标志":
                    ws.cell(row=row_idx, column=col_idx, value="0")

                if template_col in col_map:  # 如果模板列名在映射里
                    df_col = col_map[template_col]  # 找到对应的 df 列名

                    if df_col in df.columns:
                        value = row_data[df_col]
                        if template_col == "原金额":  # 模板列名为“原金额”
                            if "Exchange rate" in df.columns and "Amount" in df.columns:
                                if row_data["Exchange rate"] == 1:
                                    value = ""  # 汇率为1时置空
                                else:
                                    value = row_data["sum_yuan"]  # 否则填amount的值

                        ws.cell(row=row_idx, column=col_idx, value=value)




        output_path = os.path.join(settings.MEDIA_ROOT, "output.xlsx")
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            print("保存成功",output_path)
        except Exception as e:
            print("保存失败:", e)



    def download_file(request, filename):
        file_path = os.path.join(settings.MEDIA_ROOT, filename)
        if not os.path.exists(file_path):
            raise Http404("文件不存在")

        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


    def name_to_email(self, name):
        if name.strip().lower() == "cherry li":
            return "cherry1.li@iss-gf.com"
        else:
            return name.replace(" ", ".") + "@iss-gf.com"


