from decimal import ROUND_CEILING, Decimal

from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser,FormParser
import pandas as pd
import io,os
from openpyxl import load_workbook
import tempfile
from django.http import FileResponse,Http404
from rest_framework.views import APIView
from rest_framework import status
from django.conf import settings
from utils import mail_utils
import platform
import traceback
import numpy as np


class Iss_Fin1_solve_excel(APIView):
    parser_classes = (MultiPartParser, FormParser)



    def post(self, request):

        self.df1 = None
        self.df2 = None
        self.df3 = None
        files = request.FILES.getlist("files")

        try:
            self.read_files(files)
            res = self.solve_excel_data()

            return res
            # return Response({
            #     'message': '上传成功',
            #     'count': non_empty_count,
            #     'files': non_empty_count
            # }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                'message': f'上传失败 {e}',
                'count': 0,
                'files': 0
            }, status=status.HTTP_201_CREATED)



    def read_files(self,files):
        try:
            for file in files:

                if isinstance(file, str):
                    filename = file
                    content = None
                else:
                    filename = file.name
                    content = file.read()
                    file.seek(0)

                if "eInvoice" in file.name:
                    # self.df1 = pd.read_excel(io.BytesIO(file.read()))

                    if content:
                        self.df1 = pd.read_excel(io.BytesIO(content),usecols=["Invoice number", "Invoice type","Listing flag","Buyer name","Tax rate","Tax classification code","Amount","Preferential policy name","Zero tax rate mark","Issuing note","Exchange rate","Remarks","Buyer tax number"],
                                               )

                    else:
                        self.df1 = pd.read_excel(file)


                elif "Invoices" in file.name:

                    if content:

                        self.df2 = pd.read_csv(io.BytesIO(content), usecols=["Invoice No.", "Created By","Remarks","US Invoice Remarks","Invoice Date"])

                    else:
                        self.df2 = pd.read_csv(file, usecols=["Invoice No.", "Created By"], encoding="utf-8-sig")



                # elif file.name.startswith("Contacts"):
                #     self.df3 = pd.read_excel(io.BytesIO(content),usecols=["Foreign Name", "Tax ID"])


                if self.df1 is None or self.df1.empty:
                    print("df1（eInvoice）没有数据或读取失败")

                if self.df2 is None or self.df2.empty:
                    print("df2（Invoices）没有数据或读取失败")

                # if self.df3 is None or self.df3.empty:
                #     print("df3（Contacts）没有数据或读取失败")
        except Exception as e:
            print("读取失败:", e)
            traceback.print_exc()


    def update_row(self,row):
        if row['Invoice type'] == "02" :
            return "国际货物运输" + str(row['开票商品名称'])
        return str(row['开票商品名称'])

    def solve_excel_data(self):

        if self.df1 is None or self.df2 is None :
            return

        df1 = self.df1.drop(self.df1.index[0]).reset_index(drop=True)

        df1 = df1.fillna('')



        self.df2["email"] = self.df2["Created By"].apply(self.name_to_email)

        df2 = self.df2.fillna('')

        df1 = df1.merge(
            df2,
            left_on="Invoice number",
            right_on="Invoice No.",
            how="left"
        )
        #找到有没写备注的invocie no
        invalid_invoice = df1.loc[
            df1["Remarks_x"].astype(str).str.strip() == "",
            "Invoice No."
        ].unique()

        # 剔除这些 Invoice No.
        df1 = df1[~df1["Invoice No."].isin(invalid_invoice)]

        group_cols = ["Invoice No.", "Remarks_x"]
        df1["Amount"] = df1["Amount"].apply(lambda x: Decimal(str(x)))
        df1 = df1.groupby(
            group_cols,
            as_index=False
        ).agg({
            "Amount": "sum",
            **{col: "first" for col in df1.columns if col not in group_cols + ["Amount"]},

        })
        # df1 = df1[df1["Commodity name"].notna() & (df1["Commodity name"].str.strip() != "")]
        cols = ['Amount', 'Exchange rate']



        for col in cols:
            df1[col] = pd.to_numeric(df1[col], errors='coerce')

        df1['sum_result'] = df1.apply(
            lambda x: float(
                (
                        Decimal(str(x['Amount'])) *
                        Decimal(str(x['Exchange rate']))
                ).quantize(
                    Decimal('0.01'),
                    rounding=ROUND_CEILING
                )
            ),
            axis=1
        )
        print("merge前面",df1[df1["Invoice No."] == "INVNGBSI26003605"])
        # df3 = self.df3.fillna('')

        # df1 = df1.merge(
        #     df3,
        #     left_on="Buyer name",
        #     right_on="Foreign Name",
        #     how="left"
        # )
        # print("merge后面",df1[df1["Invoice No."] == "INVNGBSI26003605"])
        self.complete_template_data(df1)
        return self.download_file("output.xlsx")


    def complete_template_data(self, df):
        try:

            if platform.system() == 'Windows':
                # Windows 系统
                template_path = r'D:\project\excelapp\excel_project\excel_api\templates\excel_api\template.xlsx'
            else:
                # Linux / macOS
                template_path = os.path.normpath(
                    os.path.join(
                        settings.BASE_DIR,
                        'excel_api',
                        'templates',
                        'excel_api',
                        'template.xlsx'  # 注意 Linux 文件名大小写敏感
                    )
                )

            # template_path = os.path.normpath(os.path.join(settings.BASE_DIR, 'excel_api', 'templates', 'excel_api',  '模版.xlsx'))
            # template_path = r'D:\project\excelapp\excel_project\excel_api\templates\excel_api\template.xlsx'
            df["zhence"] = ''
            df["biaoshi"] = ''
            df.loc[df['Invoice type'] == '02', 'zhence'] = '免税'
            df.loc[df['Invoice type'] == '02', 'biaoshi'] = 1
            df["remark_new"] = df["Issuing note"].fillna("") +"\n" + df["Remarks_y"].fillna("")
            wb = load_workbook(template_path)

            ws = wb["普通发票"]  # 如果有多个 sheet，可以 ws = wb["Sheet1"]
            ws1 = wb["04货物运输服务"]

            template_columns = [ws.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)]
            template_columns_1 = [ws1.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)]


            start_row = 4
            row_idx_ws1 = 4
            col_map = {
                "开票单号*": "Invoice number",
                "购方名称*": "Buyer name",
                "购方税号": "Buyer tax number",
                "商品名称*": "Remarks_x",
                "单价": "sum_result",
                "金额": "sum_result",
                "发票备注": "remark_new",
                "邮箱地址": "email",
                "税率": "Tax rate",
                "发票种类*": "Invoice type",
                "零税率标识": "biaoshi",
                "优惠政策名称": "zhence",
                "原金额": "Amount",
                "US Invoice Remarks":"US Invoice Remarks",
                "Invoice Date":"Invoice Date"

            }

            col_map_1 = {
                "开票单号*": "Invoice number"
            }


            for row_idx, row_data in enumerate(df.to_dict(orient="records"), start=start_row):
                for col_idx, template_col in enumerate(template_columns, start=1):
                    if template_col == "税收分类编码":
                        ws.cell(row=row_idx, column=col_idx, value="3040802010200000000")
                    if template_col == "数量":
                        ws.cell(row=row_idx, column=col_idx, value="1.00")
                    if template_col == "清单标志":
                        ws.cell(row=row_idx, column=col_idx, value="0")

                    if template_col in col_map:  # 如果模板列名在映射里
                        df_col = col_map[template_col]  # 找到对应的 df 列名

                        if df_col in df.columns:
                            value = row_data[df_col]
                            if template_col == "原金额":  # 模板列名为“原金额”
                                if "Exchange rate" in df.columns and "Amount" in df.columns:
                                    if row_data["Exchange rate"] == 1:
                                        value = ""  # 汇率为1时置空
                                    else:
                                        value = row_data["Amount"]  # 否则填amount的值

                            ws.cell(row=row_idx, column=col_idx, value=value)


                if str(row_data.get("Tax rate", "")).strip() not in ["9%", "9","0.09"]:
                    continue
                for col_idx, template_col in enumerate(template_columns_1, start=1):
                    if template_col in col_map_1:
                        df_col = col_map_1[template_col]  # 找到对应的 df 列名
                        if df_col in df.columns:
                            value = row_data[df_col]
                            ws1.cell(row=row_idx_ws1 , column=col_idx, value=value)
                            row_idx_ws1 +=1
                            print(value)

            output_path = os.path.join(settings.MEDIA_ROOT, "output.xlsx")

            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            print("保存成功",output_path)
        except Exception as e:
            print("保存失败:", e)
            traceback.print_exc()



    def download_file(request, filename):
        file_path = os.path.join(settings.MEDIA_ROOT, filename)
        if not os.path.exists(file_path):
            raise Http404("文件不存在")

        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


    def name_to_email(self, name):
        if name.strip().lower() == "cherry li":
            return "cherry1.li@iss-gf.com"
        elif name.strip().lower() == "NGBOP.PUBLIC@iss-gf.com":
            return "queeny.wu@iss-gf.com"
        else:
            return name.replace(" ", ".") + "@iss-gf.com"


