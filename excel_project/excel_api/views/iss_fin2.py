from decimal import ROUND_CEILING, Decimal

from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser,FormParser
import pandas as pd
import io,os
from openpyxl import load_workbook
import tempfile
from django.http import FileResponse,Http404
from rest_framework.views import APIView
from rest_framework import status
from django.conf import settings
from utils import mail_utils
import platform
import traceback
import numpy as np


class Iss_Fin2_solve_excel(APIView):
    parser_classes = (MultiPartParser, FormParser)



    def post(self, request):

        self.df1 = None
        self.df2 = None
        files = request.FILES.getlist("files")

        try:
            self.read_files(files)
            res = self.solve_excel_data()
            return res
            # return Response({
            #     'message': '上传成功',
            #     'count': non_empty_count,
            #     'files': non_empty_count
            # }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                'message': f'上传失败 {e}',
                'count': 0,
                'files': 0
            }, status=status.HTTP_201_CREATED)

    def read_files(self, files):

        try:

            for file in files:

                filename = file.name

                # 读取上传内容
                content = file.read()

                # 重要：每次读取后重置指针
                file.seek(0)

                # 根据文件名区分
                if "111" in filename:

                    if platform.system() == 'Windows':
                        # Windows 系统
                        template_path = r'D:\project\excelapp\excel_project\excel_api\templates\test111.xlsx'
                    else:
                        # Linux / macOS
                        template_path = os.path.normpath(
                            os.path.join(
                                settings.BASE_DIR,
                                'excel_api',
                                'templates',
                                'test111.xlsx'  # 注意 Linux 文件名大小写敏感
                            )
                        )

                    # 保存为新模板
                    with open(template_path, "wb") as f:
                        f.write(content)

                    print("模板已更新")

                    self.df1 = pd.read_excel(
                        io.BytesIO(content),
                        sheet_name ="Overdue Billing",
                        header=2,
                        usecols=["Serial No", "YES or NO\nBillable immediately (next 24h)?","WHY\nFor NO Answers ONLY, please give a reason.","WHEN\nFor NO answers ONLY,\nprovide the billing date."]
                    )
                    print(self.df1.columns)

                    print("读取111成功")
                    # print(self.df1)
                elif "222" in filename:
                    #老版本
                    # self.df2 = pd.read_excel(
                    #     io.BytesIO(content),
                    #     sheet_name="USD with aging",
                    #     header=1,
                    #     usecols=["Sab-JobID", "If NO, we need the reason WHY","Indicate the date WHEN will you bill iit (if It is after the weekend)"]
                    # )

                    ##新版本
                    self.df2 = pd.read_excel(
                        io.BytesIO(content),
                        sheet_name="Overdue Billing",
                        header=2,
                        usecols=["Serial No", "YES or NO\nBillable immediately (next 24h)?",
                                 "WHY\nFor NO Answers ONLY, please give a reason.",
                                 "WHEN\nFor NO answers ONLY,\nprovide the billing date."]
                    )

                    print("读取222成功")
                    # print(self.df2)



                # if self.df3 is None or self.df3.empty:
                #     print("df3（Contacts）没有数据或读取失败")
        except Exception as e:
            print("读取失败:", e)
            traceback.print_exc()

    def solve_excel_data(self):

        if self.df1 is None or self.df2 is None:
            return
        # 清洗
        self.df1["Serial No"] = self.df1["Serial No"].astype(str).str.strip()

        # self.df2["Sab-JobID"] = self.df2["Sab-JobID"].astype(str).str.strip()
        self.df2["Serial No"] = self.df2["Serial No"].astype(str).str.strip()
        # 建立why映射
        # reason_map = self.df2.set_index("Sab-JobID")[
        #     "If NO, we need the reason WHY"
        # ]

        reason_map = self.df2.set_index("Serial No")[
            "WHY\nFor NO Answers ONLY, please give a reason."
        ]
        target_col = "WHY\nFor NO Answers ONLY, please give a reason."
        self.df1[target_col] = self.df1[target_col].fillna(
            self.df1["Serial No"].map(reason_map)
        )

        # =========================
        # date 映射
        # =========================
        # date_map = self.df2.set_index("Sab-JobID")[
        #     "Indicate the date WHEN will you bill iit (if It is after the weekend)"
        # ]
        date_map = self.df2.set_index("Serial No")[
            'WHEN\nFor NO answers ONLY,\nprovide the billing date.'
        ]

        # 只填空值
        self.df1['WHEN\nFor NO answers ONLY,\nprovide the billing date.'] = self.df1['WHEN\nFor NO answers ONLY,\nprovide the billing date.'].fillna(
            self.df1["Serial No"].map(date_map)
        )

        #
        self.df1["YES or NO\nBillable immediately (next 24h)?"] = self.df1["WHY\nFor NO Answers ONLY, please give a reason."].apply(
            lambda x: "NO" if pd.notna(x) and str(x).strip() != "" else ""
        )


        self.complete_template_data(self.df1)
        return self.download_file("output111.xlsx")



    def complete_template_data(self, df):
        try:

            if platform.system() == 'Windows':
                # Windows 系统
                template_path = r'D:\project\excelapp\excel_project\excel_api\templates\test111.xlsx'
            else:
                # Linux / macOS
                template_path = os.path.normpath(
                    os.path.join(
                        settings.BASE_DIR,
                        'excel_api',
                        'templates',
                        'test111.xlsx'  # 注意 Linux 文件名大小写敏感
                    )
                )



            print(template_path)
            wb = load_workbook(template_path)
            ws = wb["Overdue Billing"]
            # 第3行是列名
            header_row = 3

            # 找列号
            col_map = {}

            for cell in ws[header_row]:
                if cell.value is not None:
                    col_name = str(cell.value).strip()

                    col_map[col_name] = cell.column

            print(col_map)

            serial_col = col_map["Serial No"]


            print(df.columns)
            # # df建立映射
            df_map = df.set_index("Serial No")
            # print(df_map.head())
            # print(df_map.columns.tolist())

            # 从第4行开始写数据
            for row in range(header_row + 1, ws.max_row + 1):

                serial_no = ws.cell(
                    row=row,
                    column=serial_col
                ).value

                if serial_no is None:
                    continue

                serial_no = str(serial_no).strip()

                # df中不存在
                if serial_no not in df_map.index:
                    continue

                # df匹配行
                df_row = df_map.loc[serial_no]

                # =========================
                # 遍历Excel所有列
                # =========================
                for col_name, col_index in col_map.items():
                    if col_name == "Serial No":
                        continue
                    # df没有该列
                    if col_name not in df.columns:
                        continue

                    # df中的新值
                    new_value = df_row[col_name]

                    # NaN跳过
                    if pd.isna(new_value):
                        continue

                    # Excel当前值
                    current_value = ws.cell(
                        row=row,
                        column=col_index
                    ).value

                    # =========================
                    # 只填空值
                    # =========================
                    if current_value in [None, ""]:
                        ws.cell(
                            row=row,
                            column=col_index
                        ).value = new_value

            # 保存（不会破坏格式）
            output_path = os.path.join(settings.MEDIA_ROOT, "output111.xlsx")

            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            print("保存成功", output_path)


        except Exception as e:
            print("保存失败:", e)
            traceback.print_exc()

    def download_file(request, filename):
        file_path = os.path.join(settings.MEDIA_ROOT, filename)
        if not os.path.exists(file_path):
            raise Http404("文件不存在")

        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


