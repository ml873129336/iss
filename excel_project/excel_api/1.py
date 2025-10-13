import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook

class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Tkinter + pandas Excel处理App")
        self.df1 = None
        self.df2 = None
        self.df3 = None
        self.filter_df = pd.DataFrame()


        # 按钮
        self.btn_load = ttk.Button(root, text="加载excel", command=lambda: self.load_excel(1))
        self.btn_load.pack(pady=5)

        self.btn1_load = ttk.Button(root, text="加载csv", command=lambda: self.load_excel(2))
        self.btn1_load.pack(pady=5)

        self.btn2_load = ttk.Button(root, text="加载excel3", command=lambda: self.load_excel(3))
        self.btn2_load.pack(pady=5)


        self.btn_third = ttk.Button(root, text="solve_excel", command=self.solve_excel_data, state='disabled')
        self.btn_third.pack(pady=20)

        self.btn_third1 = ttk.Button(root, text="读取模版并处理", command=lambda: self.complete_template_data(self.filter_df), state='disabled')
        self.btn_third1.pack(pady=20)



    def load_excel(self, num):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.csv" )])
        if not path:
            return
        try:

            if num == 1:
                self.df1 = pd.read_excel(path)
                # self.df3 = pd.read_excel(path,sheet_name='Sheet1')

                messagebox.showinfo("提示", "表格1加载成功！")
            elif num== 2:
                self.df2 = pd.read_csv(path, usecols=["Invoice No.", "Created By"], encoding="utf-8-sig")
                messagebox.showinfo("提示", "表格2加载成功！")
            elif num == 3:
                sheets_to_read = ['海运部', '深圳办', '宁波办', '空运部', '项目部','税率6%']
                # sheets_to_read = ['临时']
                dfs = [pd.read_excel(path, sheet_name=s,skiprows=1,usecols=["INVPVGNo.\n不要有空格", "开票商品名称","备注"]) for s in sheets_to_read]

                self.df3 = pd.concat(dfs, ignore_index=True)
                self.df3 = self.df3.replace('_x000D_', '', regex=True)
                # self.df3 = self.df3.dropna(subset=['INVPVGNo.\n不要有空格'])

                self.df3['INVPVGNo.\n不要有空格'] = self.df3['INVPVGNo.\n不要有空格'].str.replace(r'\s+', '', regex=True)

                messagebox.showinfo("提示", "表格3加载成功！")


            self.check_both_loaded()
        except Exception as e:
            messagebox.showerror("错误", f"加载表格{num}失败：{e}")

    def check_both_loaded(self):

        if self.df1 is not None and self.df2 is not None:
            self.btn_third.config(state='normal')



    def solve_excel_data(self):
        if self.df1 is None or self.df2 is None:
            messagebox.showerror("错误", "请先加载两个表格！")
            return
        df1 = self.df1.drop(self.df1.index[0]).reset_index(drop=True)

        df = df1.fillna('')
        print("处理前df" + str(len(df1)))
        cols = ['Amount', 'Exchange rate']
        for col in cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        df['sum1'] = df['Amount'] * df['Exchange rate']
        df['sum'] = df.groupby(['Invoice number'])['sum1'].transform('sum')


        last_col_data = df.iloc[:, -1]

        df.loc[:, 'Unit price'] = last_col_data
        df.loc[:, 'Amount'] = last_col_data
        df = df.rename(columns={"Invoice number": "Invoice No."})
        df.drop(columns=['sum1'], inplace=True)

        self.df2["email"] = self.df2["Created By"].apply(self.name_to_email)

        inv_to_email = pd.Series(self.df2["email"].values, index=self.df2["Invoice No."]).to_dict()
        df["Email address"] = df["Invoice No."].map(inv_to_email)
        df = df.drop_duplicates(subset = ['Invoice No.'])
        print("处理后df"+str(len(df)))

        # print(self.df3)
        if self.df3 is not None and not self.df3.empty:
            self.btn_third1.config(state='normal')
            print(self.df3[self.df3['INVPVGNo.\n不要有空格'] == 'INVSZXSE25006950'])
            df = df[df['Invoice No.'].isin(self.df3['INVPVGNo.\n不要有空格'])]
            df_merged = df.merge(
                self.df3[['INVPVGNo.\n不要有空格', '开票商品名称', '备注']],
                left_on='Invoice No.',
                right_on='INVPVGNo.\n不要有空格',
                how='left',
            )

            df_merged['Commodity name'] = df_merged['Commodity name'].where(
                df_merged['Commodity name'] != '',
                "国际货物运输"+df_merged['开票商品名称']
            )
            df_merged['备注'] = df_merged['备注'].fillna('').astype(str)
            df_merged['Issuing note'] =df_merged['Issuing note']+"\n" + df_merged['备注']

            df=df_merged.drop(columns=['开票商品名称', '备注'])
            self.filter_df = df
            if not self.filter_df.empty:
                self.btn_third1.config(state='normal')
        self.save_to_excel(df)





    def complete_template_data(self,df):
        df["zhence"] = ''
        df["biaoshi"] = ''
        df.loc[df['Invoice type'] == '02', 'zhence'] = '免税'
        df.loc[df['Invoice type'] == '02', 'biaoshi'] = 1
        template_path = "templates/excel_api/template.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active  # 如果有多个 sheet，可以 ws = wb["Sheet1"]
        template_columns = [ws.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)]
        start_row = 4
        col_map = {
            "开票单号*": "Invoice No.",
            "购方名称*": "Buyer name",
            "购方税号": "Buyer tax number",
            "商品名称*": "Commodity name",
            "单价": "Unit price",
            "金额": "Amount",
            "发票备注": "Issuing note",
            "邮箱地址": "Email address",
            "税率": "Tax rate",
            "发票种类*":"Invoice type",
            "零税率标识":"biaoshi",
            "优惠政策名称":"zhence"

        }

        for row_idx, row_data in enumerate(df.to_dict(orient="records"), start=start_row):
            for col_idx, template_col in enumerate(template_columns, start=1):
                if template_col == "税收分类编码":
                    ws.cell(row=row_idx, column=col_idx, value="3040802010200000000")
                if template_col == "数量":
                    ws.cell(row=row_idx, column=col_idx, value="1.00")
                if template_col == "清单标志":
                    ws.cell(row=row_idx, column=col_idx, value="0")
                if template_col in col_map:  # 如果模板列名在映射里
                    df_col = col_map[template_col]  # 找到对应的 df 列名
                    if df_col in df.columns:
                        ws.cell(row=row_idx, column=col_idx, value=row_data[df_col])


        wb.save("填充后的文件.xlsx")

    def save_to_excel(self,df):

        if df is None:
            messagebox.showwarning("警告", "没有数据可保存")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            df.to_excel(file_path, index=False)
            messagebox.showinfo("成功", "保存成功！")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}")



    def name_to_email(self,name):
        if name.strip().lower() == "cherry li":
            return "cherry1.li@iss-gf.com"
        if name.strip().lower() == "NGBOP PUBLIC":
            return "queeny.wu@iss-gf.com"
        else:
            return name.replace(" ", ".") + "@iss-gf.com"





if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("600x400")
    app = ExcelApp(root)
    root.mainloop()
