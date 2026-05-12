import os
from datetime import datetime,timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from django.conf import settings
from django.http import FileResponse, Http404
from rest_framework.decorators import api_view
from rest_framework.response import Response
from utils import mail_utils  # 根据你路径调整
from openpyxl import load_workbook
import re
from utils.excel_utils import excel_bytes_to_image_base64


CACHE={}

EXCEL_PATH = os.path.join(settings.MEDIA_ROOT, "payment/dianxin.xlsx")
IMG_PATH = os.path.join(settings.MEDIA_ROOT, "payment/dianxin.png")

COMPANY_CONFIG = {
    "dianxin": {
        "template": os.path.join("asserts_manager", "templates", "dianxin.xlsx"),
        "fields": {
            "date": "B7",
            "amount": "H26",
            "remark": "H30"
        },
        "subject": "电信付款单"
    },
    "colipu": {
        "template": os.path.join("excel_api", "templates", "excel_api", "colipu.xlsx"),
        "fields": {
            "date": "B7",
            "amount": "H25"
        },
        "subject": "科力普付款单"
    }
}

def get_paths(company):
    base_dir = os.path.join(settings.MEDIA_ROOT, "payment", company)
    os.makedirs(base_dir, exist_ok=True)

    return {
        "excel": os.path.join(base_dir, f"{company}.xlsx"),
        "img": os.path.join(base_dir, f"{company}.png"),
        "amount": os.path.join(base_dir, f"{company}.txt")
    }


@api_view(['GET'])
def it_payment_colipu(request,):
    try:
        # 1️⃣ 获取邮件内容
        body_list = mail_utils.check_email("科力普")


        data_list = extract_invoice_info(body_list)

        results = []

        print(data_list["amount"])

        # 2️⃣ 金额（支持前端传）
        amount = request.query_params.get("amount", data_list["amount"])

        # 3️⃣ 模板路径
        template_path = os.path.join(
            settings.BASE_DIR,
            'excel_api',
            'templates',
            'excel_api',
            'colipu.xlsx'
        )

        if not os.path.exists(template_path):
            raise Http404("模板不存在")

        # 4️⃣ 处理 Excel

        wb = load_workbook(template_path)
        ws = wb.active

        ws['B7'] = datetime.today().strftime('%Y-%m-%d')
        ws['H25'] = float(amount)



        # ✅ 5️⃣ 不落盘，直接内存下载（推荐）
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'colipu_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'

        return FileResponse(
            output,
            as_attachment=True,
            filename=filename
        )

    except Exception as e:

        return Response({
            "status": "error",
            "msg": str(e)
        })



@api_view(['POST'])
def payment_preview(request):
    try:
        amount = Decimal(str(request.data.get('amount')))
        company = request.data.get("company")

        paths = get_paths(company)
        output = build_excel(company, amount)
        excel_bytes = output.getvalue()

        with open(paths["excel"], "wb") as f:
            f.write(excel_bytes)

        output.seek(0)

        preview = excel_bytes_to_image_base64(output)
        img_bytes = base64.b64decode(preview)

        with open(paths["img"], "wb") as f:
            f.write(img_bytes)

        with open(paths["amount"], "w") as f:
            f.write(str(amount))

        return Response({
            "status": "success",
            "preview": preview
        })

    except Exception as e:
        return Response({"status": "error", "msg": str(e)})

@api_view(['POST'])
def payment_download(request):
    try:
        company = request.GET.get("company")
        paths = get_paths(company)
        if not os.path.exists(paths["excel"]):
            return Response({"status": "error", "msg": "请先生成预览"})

        return FileResponse(
            open(paths["excel"], "rb"),
            as_attachment=True,
            filename=f"{company}.xlsx"
        )


    except Exception as e:
        return Response({"status": "error", "msg": str(e)})

import base64

@api_view(['POST'])
def payment_send_email(request):
    try:
        company = request.data.get("company")
        email = request.data.get("email")

        config = COMPANY_CONFIG.get(company)
        paths = get_paths(company)

        if not config:
            return Response({"status": "error", "msg": "不支持的公司"})
        if not email :
            return Response({"status": "error", "msg": "缺少参数"})
        if not os.path.exists(paths["excel"]):
            return Response({"status": "error", "msg": "请先生成预览"})

        with open(paths["excel"], "rb") as f:
            excel_bytes = f.read()

        with open(paths["img"], "rb") as f:
            img_bytes = f.read()

        with open(paths["amount"], "r") as f:
            amount = f.read()


        # ✅ 3. 邮件内容
        body = f"""
        <p>请查收付款单（预览如下）：</p>
        <p>金额： {amount}</p>
        <img src="cid:preview_img" style="max-width:100%;">
        """

        # ✅ 4. 发送邮件
        mail_utils.send_email(
            to=email,
            subject=config["subject"],
            body=body,
            body_type="html",
            attachments=[(f"{company}.xlsx", excel_bytes)],
            inline_image=("preview_img", img_bytes)
        )

        return Response({"status": "success", "msg": "邮件已发送"})

    except Exception as e:
        return Response({"status": "error", "msg": str(e)})

def build_excel(company,amount):
    config = COMPANY_CONFIG.get(company)
    if not config:
        raise Exception("不支持的公司")

    template_path = os.path.join(settings.BASE_DIR, config["template"])

    wb = load_workbook(template_path)
    ws = wb.active

    today = datetime.today()
    fields = config["fields"]
    if "date" in fields:
        ws[fields["date"]] = today.strftime('%Y-%m-%d')

    if "amount" in fields:
        ws[fields["amount"]] = float(amount)

    if "remark" in fields:
        first_day = today.replace(day=1)
        last_day = first_day - timedelta(days=1)
        first_day_last = last_day.replace(day=1)
        date_str = f"{first_day_last.strftime('%Y.%m.%d')}至{last_day.strftime('%Y.%m.%d')}"
        ws[fields["remark"]] = f"备注：{date_str}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)


    return output


def extract_invoice_info(body_list):
    data = {}
    for body in body_list:


        # # 发票号
        # invoice_no = re.search(r"发票号码[:：]?\s*(\d+)", body["body"])
        # if invoice_no:
        #     data["invoice_no"] = invoice_no.group(1)

        # 金额
        amount = re.search(r"合计金额[:：]?\s*([\d,]+(?:\.\d+)?)", body["body"])
        if amount:
            data["amount"] = float(amount.group(1).replace(",", ""))

        # # 出库单号
        # order_no = re.search(r"出库单号[:：]?\s*([\d；;]+)", body["body"])
        # if order_no:
        #     data["order_no"] = order_no.group(1)

    return data